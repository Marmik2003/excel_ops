from webbrowser import get
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class GetExcel:
    def __init__(self, data):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Main'
        self.data = data
        self.topics_sheet = self.wb.create_sheet(title='Topics')

    def _create_topics_sheet(self):
        max_subtopic_length = 0
        tl = len(self.data)
        self.tl = tl
        topics = [i['name'] for i in self.data]
        self.topics_sheet.append(topics)
        for idx, topic in enumerate(self.data):
            subtopics = [i['name'] for i in topic['subtopics']]
            for idx2, subtopic in enumerate(subtopics):
                self.topics_sheet.cell(row=idx2 + 2, column=idx + 1).value = subtopic
            if len(subtopics) > max_subtopic_length:
                max_subtopic_length = len(subtopic)
        self.max_subtopic_length = max_subtopic_length
        self.topics_sheet.sheet_state = 'veryHidden'

    def create(self):
        self._create_topics_sheet()
        self.ws.append(
            ['Topic', 'Subtopic', 'Question', 'Type of Answer', 'Number of options'] +
            ['Option {}'.format(i) for i in range(1, 9)] +
            ['Answer',  'Explanation', 'Difficulty', 'Marks', 'Negative marks']
        )
        topic_dv = DataValidation(
            type="list", formula1=f'\'Topics\'!$A$1:${get_column_letter(self.tl)}$1', allow_blank=False
        )
        self.ws.add_data_validation(topic_dv)
        topic_dv.add('A2:A5000')
        max_ltr = get_column_letter(self.tl)
        for i in range(2, 5001):
            subtopic_dv = DataValidation(
                type="list",
                formula1=f"=OFFSET(Topics!A1,1,MATCH(Main!A{i},Topics!A1:{max_ltr}1,0)-1,{self.max_subtopic_length},1)"
            )
            self.ws.add_data_validation(subtopic_dv)
            subtopic_dv.add(f'B{i}:B{i}')

        anstype_dv = DataValidation(
            type="list",
            formula1='"Single, Multiple"',
            allow_blank=False
        )
        self.ws.add_data_validation(anstype_dv)
        anstype_dv.add('D2:D5000')

        numopts_dv = DataValidation(
            type="whole",
            formula1='1:8',
            allow_blank=False
        )
        self.ws.add_data_validation(numopts_dv)
        numopts_dv.add('E2:E5000')

        answer_dv = DataValidation(
            type="list",
            formula1='"{}"'.format(','.join(["Option " + str(i) for i in range(1, 9)])),
            allow_blank=False
        )
        self.ws.add_data_validation(answer_dv)
        answer_dv.add('N2:N5000')

        difficulty_dv = DataValidation(
            type="list",
            formula1='"Easy, Medium, Hard"',
            allow_blank=False
        )
        self.ws.add_data_validation(difficulty_dv)
        difficulty_dv.add('P2:P5000')

        for i in range(1, 20):
            self.ws.column_dimensions[get_column_letter(i)].width = 30

        return self.wb
